# -*- coding: utf-8 -*-
"""
Created on Sat Sep 17 13:15:27 2022

@author: shintaro
"""

###import library
import datetime
import json
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import os
from pathlib import Path
import PIL
import requests

###variable

##time
now = datetime.datetime.now()

##dir_path
app_dir = Path.cwd().parent
dat_dir = app_dir / "dat"
image_dir = app_dir / "image"
tmp_dir = app_dir / "tmp"
etc_dir = app_dir / "etc"

##output_file
excel_file_name = 'output_' + now.strftime('%Y%m%d%H%M') + '.xlsx'
excel_file_path = dat_dir / excel_file_name

##temp_file
target_json = 'target_data.json'
target_json = tmp_dir / target_json

##env
ENV_FILE = "property.json"

def main():
    print("start app!")
    init()
    
    #search
    hash_id = get_hash_id(instagram_id, access_token, query)
    data = get_hash_search_result(hash_id, instagram_id, access_token)
    print("finish hashtag search!")
    
    #write_intermediate_file
    write_json_file(data,target_json)
    
    #download_image
    post_informations = download_image_and_get_post_list(data["data"], image_dir)
    print("finish download!")
    
    #write_output_file
    write_image_and_captions(post_informations,image_dir, excel_file_path)
    print("finish make output file")
    
    print("finish app!")

def init():
    global instagram_id
    global access_token
    global query
    
    files = os.listdir(image_dir)

    for file in files:
        os.remove(str(image_dir / file))
        
    with open(str(etc_dir / ENV_FILE), 'r', encoding='utf-8') as f:
        json_load = json.load(f)
        
        instagram_id = json_load["access_env"][0]["instagram_id"]
        access_token = json_load["access_env"][0]["access_token"]
        query = json_load["search_env"][0]["query"]

def get_hash_id(instagram_id:str, access_token:str, query:str) -> str:
    
    id_search_url = "https://graph.facebook.com/ig_hashtag_search?user_id=" \
        + instagram_id + "&q=" + query +  "&access_token=" + access_token
    
    response = requests.get(id_search_url)
    hash_id = response.json()['data'][0]['id']
    
    return hash_id

def get_hash_search_result(hash_id:str, instagram_id:str, access_token:str) -> dict:
    
    image_search_url = "https://graph.facebook.com/" + hash_id + \
        "/recent_media?user_id=" + instagram_id \
            + "&fields=id,media_type,media_url,children{media_url,id},caption,like_count,permalink"\
            +  "&access_token=" + access_token
    
    response = requests.get(image_search_url)
    data = response.json()
    
    return data

def write_json_file(json_data:dict, file_path:str):
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)

def download_image_and_get_post_list(data:list, dir_path:str) -> list:

    captions = []
    images = []
    
    for i in data:
        try:
            if i["media_type"] == "CAROUSEL_ALBUM":
                for j in i["children"]["data"]:
                    file_name = '{}.jpg'.format(j['id'])
                    file_path = os.path.join(dir_path,file_name)
                    
                    response = requests.get(j['media_url'])
                    image = response.content
                    
                    with open(file_path, "wb") as f:
                        f.write(image)
                    
                    images.append(file_name)
                    captions.append(i["caption"])
                                        
            else:
                file_name = '{}.jpg'.format(i['id'])
                file_path = os.path.join(dir_path,file_name)
                
                response = requests.get(i['media_url'])
                image = response.content
                
                with open(file_path, "wb") as f:
                    f.write(image)
                
                images.append(file_name)
                captions.append(i["caption"])
                        
        except KeyError:
            pass
    
    return [images,captions]

def write_image_and_captions(post_informations:str, dir_path:str, output_excel_path:str):
    
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    
    now_cell = 1

    for file, caption in zip(post_informations[0],post_informations[1]):
        try:
            ws.row_dimensions[now_cell].height = 120
            
            path = os.path.join(dir_path,file)
            image = Image(path)
            
            image.height = 150
            image.width = 200
            
            ws.add_image(image,"C" + str(now_cell))
            
            ws["B" + str(now_cell)] = caption
            ws["B" + str(now_cell)].alignment =  Alignment(wrap_text=True)
                    
            now_cell+=1
        except PIL.UnidentifiedImageError:
            os.remove(path)
    
    wb.save(output_excel_path)
    wb.close()

if __name__ == "__main__":
    main()